import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import streamlit as st
import pandas as pd
from dateutil.relativedelta import relativedelta
import requests
from io import BytesIO
import xlsxwriter
from dotenv import load_dotenv

# --- إعدادات الصفحة ---
st.set_page_config(page_title="Soufy Insurance Suite", layout="wide")


# ─────────────────────────────────────────────
# دوال Telegram
# ─────────────────────────────────────────────
def send_telegram_file(file_bytes, file_name):
    token = st.secrets["TELEGRAM_TOKEN"]
    chat_id = st.secrets["TELEGRAM_CHAT_ID"]
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    files = {'document': (file_name, file_bytes)}
    data = {'chat_id': chat_id}
    try:
        requests.post(url, data=data, files=files, timeout=10)
    except:
        pass


load_dotenv()


def send_telegram_report(message):
    token = st.secrets["TELEGRAM_TOKEN"]
    chat_id = st.secrets["TELEGRAM_CHAT_ID"]
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
    try:
        requests.post(url, data=payload, timeout=5)
    except:
        pass


# ─────────────────────────────────────────────
# دوال الاتصال بـ Google Sheet
# ─────────────────────────────────────────────
def get_export_url(sheet_url):
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', sheet_url)
    if not match:
        return None
    doc_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"


@st.cache_resource(ttl=300)
def load_data(url):
    try:
        export_url = get_export_url(url)
        if not export_url:
            st.error("الرابط غير صالح")
            return None
        response = requests.get(export_url, timeout=15)
        response.raise_for_status()
        return pd.ExcelFile(BytesIO(response.content), engine='openpyxl')
    except Exception as e:
        st.error(f"فشل الاتصال: {e}")
        return None


# ─────────────────────────────────────────────
# الهيكل الثابت للمقارنة
# ─────────────────────────────────────────────
STRUCTURE = {
    "Life Insurance": "تأمين على الحياة",
    "Accidental Death": "الوفاة نتيجة حادث",
    "Total Permanent Disability": "عجز كلي دائم",
    "Partial Permanent Disability": "عجز جزئي دائم",
    "Annual Ceiling Per Person": "الحد السنوي للفرد",
    "Accommodation": "الإقامة",
    "Medical TPA": "إدارة المطالبات الطبية (TPA)",
    "Network": "الشبكة الطبية",
    "Catigorization Availability": "توفر التصنيف",
    "Inpatient": "الإقامة الداخلية (العمليات)",
    "Outpatient": "العيادات الخارجية",  # فاصل
    "Outpatient Consultations": "استشارات العيادات الخارجية",
    "Labs Scans": "تحاليل وأشعة",
    "Labs Scans Inside Hospital": "تحاليل وأشعة داخل المستشفى",
    "Physiotherapy": "العلاج الطبيعي",
    "Medication": "الأدوية",
    "Extra Benefits": "مزايا إضافية",  # فاصل
    "Dental Coverage": "تغطية الأسنان",
    "Dental Population": "عدد المستفيدين من تغطية الأسنان",
    "Dental Limit": "حد تغطية الأسنان",
    "Optical Coverage": "تغطية النظارات (البصرية)",
    "Optical Population": "عدد المستفيدين من تغطية النظارات",
    "Optical Limit": "حد تغطية النظارات",
    "Maternity Limit": "حد تغطية الولادة",
    "Maternity Population": "عدد المستفيدين من تغطية الولادة",
    "Waiting Period": "فترة الانتظار",
    "New Born Ceiling": "حد تغطية المولود الجديد",
    "Chronic And Pre-existing": "الأمراض المزمنة والحالات السابقة",
    "New Chronic": "الأمراض المزمنة الجديدة",
    "Reimbursement Coverage": "تغطية التعويض النقدي",
    "Reimbursement Doctor Visit Coverage Up To": "تغطية زيارات الطبيب حتى",
    "Reimbursement Hospitals": "المستشفيات المعتمدة للتعويض"
}


# ─────────────────────────────────────────────
# الدوال الحسابية
# ─────────────────────────────────────────────
def calculate_insurance_age(birth_date, calculation_date=None, company_name=""):
    if pd.isna(birth_date):
        return 0
    calc = pd.to_datetime(calculation_date) if calculation_date else datetime.today()
    diff = relativedelta(calc, birth_date)
    age_years, age_months = diff.years, diff.months
    if any(n in str(company_name).lower() for n in ["sarwa", "ثروه"]):
        return age_years + 1 if age_months >= 6 else age_years
    elif any(n in str(company_name).lower() for n in ["kaf", "كاف"]):
        return age_years + 1 if age_months >= 9 else age_years
    elif any(n in str(company_name).lower() for n in ["libanosuisse", "ليبانو"]):
        return age_years + 1 if age_months >= 5.5 else age_years
    return age_years


def get_price(age, price_list, selected_column_name):
    for _, row in price_list.iloc[1:].iterrows():
        try:
            raw_range = str(row['Plan Name']).replace(" ", "")
            low, high = map(int, raw_range.split('-')) if '-' in raw_range else (int(raw_range), int(raw_range))
            if low <= age <= high:
                val = row[selected_column_name]
                return float(val) if pd.notna(val) and str(val).upper() != 'NA' else 0
        except:
            continue
    return 0


# ─────────────────────────────────────────────
# دالة قراءة بيانات خطة من الشيت
# ─────────────────────────────────────────────
def _get_plan_column_values(xl, company, plan):
    """تقرأ قيم الهيكل الكامل لخطة معينة وترجعها كـ list."""
    df = pd.read_excel(xl, sheet_name=company, engine='openpyxl')
    df.columns = [str(c).strip() for c in df.columns]

    plan_names_raw = df.iloc[0, 1:].tolist()
    plan_names = []
    for p in plan_names_raw:
        if pd.isna(p):
            plan_names.append("")
        else:
            p_str = str(p).strip()
            if p_str.endswith('.0'):
                p_str = p_str[:-2]
            plan_names.append(p_str)

    try:
        col_idx = plan_names.index(plan) + 1
    except ValueError:
        col_idx = 1

    col_vals = []
    for key in STRUCTURE.keys():
        if key in ["Outpatient", "Extra Benefits"]:
            col_vals.append("SECTION")
        else:
            mask = df.iloc[:, 0].astype(str).str.strip().str.lower() == key.strip().lower()
            if mask.any():
                val = df.loc[mask, df.columns[col_idx]].values[0]
            else:
                val = "-"
            col_vals.append(val)
    return col_vals


# ─────────────────────────────────────────────
# دالة إنشاء Excel العادية (بدون Categorization)
# ─────────────────────────────────────────────
def _apply_outer_frame(ws, start_row, end_row, start_col, end_col, color="1E3A5F", style="medium"):
    frame_side = Side(style=style, color=color)

    for row_idx in range(start_row, end_row + 1):
        left_cell = ws.cell(row_idx, start_col)
        left_cell.border = Border(
            left=frame_side,
            right=left_cell.border.right,
            top=left_cell.border.top,
            bottom=left_cell.border.bottom
        )

        right_cell = ws.cell(row_idx, end_col)
        right_cell.border = Border(
            left=right_cell.border.left,
            right=frame_side,
            top=right_cell.border.top,
            bottom=right_cell.border.bottom
        )

    for col_idx in range(start_col, end_col + 1):
        top_cell = ws.cell(start_row, col_idx)
        top_cell.border = Border(
            left=top_cell.border.left,
            right=top_cell.border.right,
            top=frame_side,
            bottom=top_cell.border.bottom
        )

        bottom_cell = ws.cell(end_row, col_idx)
        bottom_cell.border = Border(
            left=bottom_cell.border.left,
            right=bottom_cell.border.right,
            top=bottom_cell.border.top,
            bottom=frame_side
        )


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Ø¯Ø§Ù„Ø© Ø¥Ù†Ø´Ø§Ø¡ Excel Ø§Ù„Ø¹Ø§Ø¯ÙŠØ© (Ø¨Ø¯ÙˆÙ† Categorization)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def generate_advanced_excel(xl, selected_plans, password="123"):
    headers = [(item["company"], item["plan"]) for item in selected_plans]
    total_cols = len(headers) + 2

    selected_data = []
    for comp, plan in headers:
        selected_data.append(_get_plan_column_values(xl, comp, plan))

    wb = Workbook()
    ws = wb.active
    ws.title = "Insurance Comparison"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    sig_cell = ws.cell(1, 1, "Prepared by: Ahmed Soufy - Corporate Medical Account Manager")
    sig_cell.font = Font(size=14, bold=True, color="1E3A5F")
    sig_cell.alignment = Alignment(horizontal="center", vertical="center")
    sig_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    date_cell = ws.cell(2, 1, f"Issued Date: {datetime.now().strftime('%Y-%m-%d')}")
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    date_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.row_dimensions[3].height = 5

    header_row = 4
    dark_blue = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    light_blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    section_blue = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    ws.cell(header_row, 1, "Insurance Company").fill = dark_blue
    ws.cell(header_row, 1).font = white_font
    for i, (comp, _) in enumerate(headers):
        cell = ws.cell(header_row, i + 2, comp)
        cell.fill = dark_blue
        cell.font = white_font
    ws.cell(header_row, total_cols, "شركة التأمين").fill = dark_blue
    ws.cell(header_row, total_cols).font = white_font

    ws.cell(header_row + 1, 1, "Plan Name").fill = light_blue
    for i, (_, plan) in enumerate(headers):
        ws.cell(header_row + 1, i + 2, plan).fill = light_blue
    ws.cell(header_row + 1, total_cols, "اسم الخطة").fill = light_blue

    curr_row = header_row + 2
    for i, (key, arabic) in enumerate(STRUCTURE.items()):
        is_section = key in ["Outpatient", "Extra Benefits"]
        ws.cell(curr_row, 1, key)
        ws.cell(curr_row, total_cols, arabic)
        if is_section:
            ws.cell(curr_row, 1).fill = section_blue
            ws.cell(curr_row, total_cols).fill = section_blue
            ws.cell(curr_row, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(curr_row, total_cols).font = Font(bold=True, color="FFFFFF")
        for col_idx, col_vals in enumerate(selected_data):
            val = col_vals[i]
            cell = ws.cell(curr_row, col_idx + 2, "" if val == "SECTION" else val)
            if is_section:
                cell.fill = section_blue
        curr_row += 1

    for label, data_key, arabic in [("Member Count", "count", "عدد الأعضاء"), ("Grand Total", "total", "الإجمالي الكلي")]:
        ws.cell(curr_row, 1, label).fill = green_fill
        ws.cell(curr_row, total_cols, arabic).fill = green_fill
        for i, item in enumerate(selected_plans):
            ws.cell(curr_row, i + 2, item[data_key]).fill = green_fill
        curr_row += 1

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=header_row, max_row=curr_row - 1, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _apply_outer_frame(ws, header_row, curr_row - 1, 1, total_cols)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

    ws.protection.password = password
    ws.protection.sheet = True
    ws.protection.enable()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Ø¯Ø§Ù„Ø© Ø¥Ù†Ø´Ø§Ø¡ Excel Ù…Ø¹ Categorization (Ø¯Ù…Ø¬ Ø§Ù„Ø´Ø±ÙƒØ© ÙˆØ§Ù„Ø³Ø¹Ø±)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def generate_categorized_excel(xl, categorized_groups, password="123"):
    all_plan_cols = []
    for g_idx, grp in enumerate(categorized_groups):
        for p in grp["plans"]:
            all_plan_cols.append({
                "company": grp["company"],
                "plan": p["plan"],
                "count": p["count"],
                "total": grp["total_price"],
                "g_idx": g_idx,
            })

    total_cols = len(all_plan_cols) + 2

    all_col_vals = []
    for pc in all_plan_cols:
        all_col_vals.append(_get_plan_column_values(xl, pc["company"], pc["plan"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Insurance Comparison"

    dark_blue = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    light_blue = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    section_blue = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    merge_price = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border_left = Side(style='medium')
    thick_border_right = Side(style='medium')

    def make_group_border(left=False, right=False):
        return Border(
            left=thick_border_left if left else Side(style='thin'),
            right=thick_border_right if right else Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    sig = ws.cell(1, 1, "Prepared by: Ahmed Soufy - Corporate Medical Account Manager")
    sig.font = Font(size=14, bold=True, color="1E3A5F")
    sig.alignment = Alignment(horizontal="center", vertical="center")
    sig.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    date_c = ws.cell(2, 1, f"Issued Date: {datetime.now().strftime('%Y-%m-%d')}")
    date_c.font = Font(size=10, italic=True)
    date_c.alignment = Alignment(horizontal="center", vertical="center")
    date_c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.row_dimensions[3].height = 5

    header_row = 4
    plan_row = 5

    ws.cell(header_row, 1, "Insurance Company").fill = dark_blue
    ws.cell(header_row, 1).font = white_font
    ws.cell(header_row, 1).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(header_row, total_cols, "شركة التأمين").fill = dark_blue
    ws.cell(header_row, total_cols).font = white_font
    ws.cell(header_row, total_cols).alignment = Alignment(horizontal="center", vertical="center")

    group_ranges = {}
    for ci, pc in enumerate(all_plan_cols):
        g_idx = pc["g_idx"]
        col_1based = ci + 2
        if g_idx not in group_ranges:
            group_ranges[g_idx] = [col_1based, col_1based]
        else:
            group_ranges[g_idx][1] = col_1based

    for g_idx, (start_c, end_c) in group_ranges.items():
        comp_name = categorized_groups[g_idx]["company"]
        if start_c == end_c:
            cell = ws.cell(header_row, start_c, comp_name)
        else:
            ws.merge_cells(
                start_row=header_row, start_column=start_c,
                end_row=header_row, end_column=end_c
            )
            cell = ws.cell(header_row, start_c, comp_name)
        cell.fill = dark_blue
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_group_border(left=True, right=True)

    ws.cell(plan_row, 1, "Plan Name").fill = light_blue
    ws.cell(plan_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(plan_row, total_cols, "اسم الخطة").fill = light_blue
    ws.cell(plan_row, total_cols).alignment = Alignment(horizontal="center", vertical="center")

    for ci, pc in enumerate(all_plan_cols):
        cell = ws.cell(plan_row, ci + 2, pc["plan"])
        cell.fill = light_blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    curr_row = plan_row + 1
    for i, (key, arabic) in enumerate(STRUCTURE.items()):
        is_section = key in ["Outpatient", "Extra Benefits"]

        ws.cell(curr_row, 1, key)
        ws.cell(curr_row, total_cols, arabic)

        if is_section:
            ws.cell(curr_row, 1).fill = section_blue
            ws.cell(curr_row, 1).font = Font(bold=True, color="FFFFFF")
            ws.cell(curr_row, total_cols).fill = section_blue
            ws.cell(curr_row, total_cols).font = Font(bold=True, color="FFFFFF")

        for ci, col_vals in enumerate(all_col_vals):
            val = col_vals[i]
            cell = ws.cell(curr_row, ci + 2, "" if val == "SECTION" else val)
            if is_section:
                cell.fill = section_blue

        curr_row += 1

    ws.cell(curr_row, 1, "Member Count").fill = green_fill
    ws.cell(curr_row, total_cols, "عدد الأعضاء").fill = green_fill
    for ci, pc in enumerate(all_plan_cols):
        cell = ws.cell(curr_row, ci + 2, pc["count"])
        cell.fill = green_fill
    curr_row += 1

    ws.cell(curr_row, 1, "Grand Total").fill = merge_price
    ws.cell(curr_row, 1).font = Font(bold=True)
    ws.cell(curr_row, total_cols, "الإجمالي الكلي").fill = merge_price
    ws.cell(curr_row, total_cols).font = Font(bold=True)

    for g_idx, (start_c, end_c) in group_ranges.items():
        price_val = categorized_groups[g_idx]["total_price"]
        if start_c == end_c:
            cell = ws.cell(curr_row, start_c, price_val)
        else:
            ws.merge_cells(
                start_row=curr_row, start_column=start_c,
                end_row=curr_row, end_column=end_c
            )
            cell = ws.cell(curr_row, start_c, price_val)
        cell.fill = merge_price
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    curr_row += 1

    for row in ws.iter_rows(min_row=header_row, max_row=curr_row - 1, min_col=1, max_col=total_cols):
        for cell in row:
            if cell.border == Border():
                cell.border = thin_border
            if not cell.alignment or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for _, (start_c, end_c) in group_ranges.items():
        for row_idx in range(header_row, curr_row):
            left_cell = ws.cell(row_idx, start_c)
            left_cell.border = Border(
                left=thick_border_left,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom
            )

            right_cell = ws.cell(row_idx, end_c)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=thick_border_right,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom
            )

    _apply_outer_frame(ws, header_row, curr_row - 1, 1, total_cols)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

    ws.protection.password = password
    ws.protection.sheet = True
    ws.protection.enable()

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

st.title("🛡️ Ahmed Soufy - Insurance Suite v0.2")

if 'selected_list' not in st.session_state:
    st.session_state['selected_list'] = []
if 'xl_data' not in st.session_state:
    st.session_state['xl_data'] = None
# state لمجموعات الـ Categorization
if 'cat_groups' not in st.session_state:
    st.session_state['cat_groups'] = []   # list of {"company", "total_price", "plans":[{"plan","count"}]}

tab1, tab2, tab3 = st.tabs(["📊 Comparison Tool", "📝 Quotation System", "💡 New Ideas"])

# ════════════════════════════════════════════
# TAB 1 – Comparison Tool
# ════════════════════════════════════════════
with tab1:
    st.markdown("### Market Comparison")

    url = "https://docs.google.com/spreadsheets/d/1yFgtBlDfK6wOdWrRwbJDGGp2anwZErZ4vS50LXLWtPc/edit"

    if 'xl_data' not in st.session_state or st.session_state['xl_data'] is None:
        with st.spinner("Connecting to Insurance Server..."):
            xl = load_data(url)
            if xl:
                st.session_state['xl_data'] = xl
                st.toast("✅ Data Synced from Server!")
            else:
                st.error("Connection Failed")

    if st.button("Sync Data 🔄"):
        with st.spinner('Loading...'):
            xl = load_data(url)
            if xl:
                st.session_state['xl_data'] = xl
                st.toast("✅ Connected to Server")

    if st.session_state['xl_data']:
        xl = st.session_state['xl_data']

        use_categorization = st.checkbox(
            "🗂️ Categorization — add multiple plans for the same company with one total price",
            value=False,
            help="Enable this when one company has multiple plans that should share one grouped total price."
        )

        st.divider()

        if not use_categorization:
            col1, col2 = st.columns(2)
            with col1:
                comp = st.selectbox("Select Company", xl.sheet_names, key="comp_select")
            with col2:
                df_comp = pd.read_excel(xl, sheet_name=comp, engine='openpyxl')
                raw_plans = df_comp.iloc[0, 1:].tolist()
                plans = []
                for p in raw_plans:
                    if pd.isna(p):
                        continue
                    p_str = str(p).strip()
                    if p_str.endswith('.0'):
                        p_str = p_str[:-2]
                    plans.append(p_str)
                plan = st.selectbox("Select Plan", plans, key="plan_select")

            col3, col4 = st.columns(2)
            with col3:
                count = st.number_input(
                    "Count",
                    min_value=1,
                    value=1,
                    step=1,
                    help="Enter the number of members for this plan."
                )
            with col4:
                price = st.number_input(
                    "Total Price",
                    min_value=0.0,
                    value=0.0,
                    step=100.0,
                    help="Enter the total price for this selected plan."
                )

            if st.button("Add to Comparison List", key="add_normal"):
                st.session_state['selected_list'].append({
                    "company": comp,
                    "plan": plan,
                    "count": count,
                    "total": price,
                })
                st.success("Item added.")

            if st.session_state['selected_list']:
                st.write("---")

                display_rows = []
                for idx, item in enumerate(st.session_state['selected_list'], start=1):
                    display_rows.append({
                        "#": idx,
                        "Company": item["company"],
                        "Plan": item["plan"],
                        "Count": item["count"],
                        "Total Price": item["total"],
                        "Remove": False,
                    })
                normal_editor_df = st.data_editor(
                    pd.DataFrame(display_rows),
                    use_container_width=True,
                    hide_index=True,
                    disabled=["#", "Company", "Plan"],
                    column_config={
                        "Count": st.column_config.NumberColumn("Count", min_value=1, step=1, format="%d"),
                        "Total Price": st.column_config.NumberColumn("Total Price", min_value=0.0, step=100.0, format="%.2f"),
                        "Remove": st.column_config.CheckboxColumn("Remove", help="Select any plan you want to delete."),
                    },
                    key="comparison_remove_editor"
                )
                remove_indices = [
                    row["#"] - 1
                    for _, row in normal_editor_df.iterrows()
                    if row["Remove"]
                ]
                if remove_indices:
                    st.session_state['selected_list'] = [
                        item for idx, item in enumerate(st.session_state['selected_list'])
                        if idx not in remove_indices
                    ]
                    st.rerun()

                normal_updated = False
                for idx, row in normal_editor_df.iterrows():
                    new_count = int(row["Count"])
                    new_total = float(row["Total Price"])
                    if (
                        st.session_state['selected_list'][idx]["count"] != new_count
                        or float(st.session_state['selected_list'][idx]["total"]) != new_total
                    ):
                        st.session_state['selected_list'][idx]["count"] = new_count
                        st.session_state['selected_list'][idx]["total"] = new_total
                        normal_updated = True
                if normal_updated:
                    st.rerun()

                if st.button("Clear List", key="clear_normal"):
                    st.session_state['selected_list'] = []
                    st.rerun()

                excel_data = generate_advanced_excel(
                    xl,
                    st.session_state['selected_list'],
                    password="123"
                )

                if st.download_button(
                    label="📥 Download Advanced Excel Report",
                    data=excel_data,
                    file_name="Insurance_Comparison_Advanced.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                ):
                    comparison_details = ""
                    total_all = 0
                    for i, item in enumerate(st.session_state['selected_list'], 1):
                        comparison_details += (
                            f"{i}. 🏢 <b>{item['company']}</b> | 📋 {item['plan']} "
                            f"| 👥 {item['count']} | 💰 {item['total']:,.2f}\n"
                        )
                        total_all += item['total']

                    report_msg = (
                        f"📊 <b>Comparison Report Downloaded</b>\n"
                        f"--------------------------\n"
                        f"{comparison_details}"
                        f"--------------------------\n"
                        f"💵 <b>Grand Total:</b> {total_all:,.2f} EGP"
                    )
                    send_telegram_report(report_msg)
                    send_telegram_file(excel_data, "Market_Comparison_Report.xlsx")

        else:
            st.markdown("#### Add Categorization Group")

            with st.container(border=True):
                cat_col1, cat_col2 = st.columns(2)
                with cat_col1:
                    cat_comp = st.selectbox("Select Company", xl.sheet_names, key="cat_comp")

                df_cat = pd.read_excel(xl, sheet_name=cat_comp, engine='openpyxl')
                raw_plans_cat = df_cat.iloc[0, 1:].tolist()
                plans_cat = []
                for p in raw_plans_cat:
                    if pd.isna(p):
                        continue
                    p_str = str(p).strip()
                    if p_str.endswith('.0'):
                        p_str = p_str[:-2]
                    plans_cat.append(p_str)

                with cat_col2:
                    selected_cat_plans = st.multiselect(
                        "Select Plans",
                        options=plans_cat,
                        key="cat_plans_select",
                        help="Select every plan that should be included in this categorization group."
                    )

                plan_counts_cat = {}
                if selected_cat_plans:
                    st.markdown("**Member Count per Plan:**")
                    count_cols = st.columns(min(len(selected_cat_plans), 4))
                    for idx, p_name in enumerate(selected_cat_plans):
                        with count_cols[idx % 4]:
                            plan_counts_cat[p_name] = st.number_input(
                                f"👥 {p_name}",
                                min_value=1,
                                value=1,
                                step=1,
                                key=f"cat_count_{p_name}"
                            )

                cat_total_price = st.number_input(
                    "Grand Total Price",
                    min_value=0.0,
                    value=0.0,
                    step=100.0,
                    key="cat_total_price",
                    help="Enter the one total price that should apply to the full categorization group."
                )

                add_cat_btn = st.button(
                    "Add Categorization Group",
                    key="add_cat_group",
                    disabled=(len(selected_cat_plans) == 0)
                )

                if add_cat_btn and selected_cat_plans:
                    st.session_state['cat_groups'].append({
                        "company": cat_comp,
                        "total_price": cat_total_price,
                        "plans": [
                            {"plan": p, "count": plan_counts_cat.get(p, 1)}
                            for p in selected_cat_plans
                        ]
                    })
                    st.success(f"Group added for {cat_comp}.")
                    st.rerun()

            if st.session_state['cat_groups']:
                st.write("---")
                st.markdown("#### Added Groups")

                summary_rows = []
                for g_idx, grp in enumerate(st.session_state['cat_groups']):
                    for p_idx, p in enumerate(grp["plans"]):
                        summary_rows.append({
                            "Group #": g_idx + 1,
                            "Company": grp["company"],
                            "Plan": p["plan"],
                            "Count": p["count"],
                            "Group Total Price": grp["total_price"],
                            "Price Row": p_idx == 0,
                            "Remove Plan": False,
                            "Remove Group": False,
                        })
                cat_editor_df = st.data_editor(
                    pd.DataFrame(summary_rows),
                    use_container_width=True,
                    hide_index=True,
                    disabled=["Group #", "Company", "Plan", "Price Row"],
                    column_config={
                        "Count": st.column_config.NumberColumn("Count", min_value=1, step=1, format="%d"),
                        "Group Total Price": st.column_config.NumberColumn("Group Total Price", min_value=0.0, step=100.0, format="%.2f"),
                        "Price Row": None,
                        "Remove Plan": st.column_config.CheckboxColumn("Remove Plan", help="Delete only the selected plan."),
                        "Remove Group": st.column_config.CheckboxColumn("Remove Group", help="Delete the full company group."),
                    },
                    key="categorization_remove_editor"
                )

                groups_to_remove = sorted(
                    {
                        int(row["Group #"]) - 1
                        for _, row in cat_editor_df.iterrows()
                        if row["Remove Group"]
                    },
                    reverse=True,
                )
                if groups_to_remove:
                    for group_idx in groups_to_remove:
                        if 0 <= group_idx < len(st.session_state['cat_groups']):
                            st.session_state['cat_groups'].pop(group_idx)
                    st.rerun()

                plans_to_remove = []
                for _, row in cat_editor_df.iterrows():
                    if row["Remove Plan"]:
                        plans_to_remove.append((int(row["Group #"]) - 1, row["Plan"]))

                if plans_to_remove:
                    for group_idx, plan_name in plans_to_remove:
                        if 0 <= group_idx < len(st.session_state['cat_groups']):
                            st.session_state['cat_groups'][group_idx]["plans"] = [
                                plan_item
                                for plan_item in st.session_state['cat_groups'][group_idx]["plans"]
                                if plan_item["plan"] != plan_name
                            ]
                    st.session_state['cat_groups'] = [
                        grp for grp in st.session_state['cat_groups']
                        if grp["plans"]
                    ]
                    st.rerun()

                cat_updated = False
                for _, row in cat_editor_df.iterrows():
                    group_idx = int(row["Group #"]) - 1
                    if not (0 <= group_idx < len(st.session_state['cat_groups'])):
                        continue

                    new_count = int(row["Count"])
                    new_group_total = float(row["Group Total Price"])
                    plan_name = row["Plan"]

                    if bool(row["Price Row"]) and float(st.session_state['cat_groups'][group_idx]["total_price"]) != new_group_total:
                        st.session_state['cat_groups'][group_idx]["total_price"] = new_group_total
                        cat_updated = True

                    for plan_item in st.session_state['cat_groups'][group_idx]["plans"]:
                        if plan_item["plan"] == plan_name and plan_item["count"] != new_count:
                            plan_item["count"] = new_count
                            cat_updated = True

                if cat_updated:
                    st.rerun()

                col_clear_cat, col_dl_cat = st.columns(2)

                with col_clear_cat:
                    if st.button("Clear All Groups", key="clear_cat"):
                        st.session_state['cat_groups'] = []
                        st.rerun()

                with col_dl_cat:
                    cat_excel_data = generate_categorized_excel(
                        xl,
                        st.session_state['cat_groups'],
                        password="123"
                    )

                    if st.download_button(
                        label="📥 Download Categorized Excel Report",
                        data=cat_excel_data,
                        file_name="Insurance_Comparison_Categorized.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    ):
                        details = ""
                        grand = 0
                        for g_idx, grp in enumerate(st.session_state['cat_groups'], 1):
                            plans_str = ", ".join([f"{p['plan']} ({p['count']})" for p in grp["plans"]])
                            details += (
                                f"{g_idx}. 🏢 <b>{grp['company']}</b>\n"
                                f"   📋 {plans_str}\n"
                                f"   💰 {grp['total_price']:,.2f} EGP\n"
                            )
                            grand += grp["total_price"]

                        report_msg = (
                            f"📊 <b>Categorized Comparison Report Downloaded</b>\n"
                            f"--------------------------\n"
                            f"{details}"
                            f"--------------------------\n"
                            f"💵 <b>Grand Total:</b> {grand:,.2f} EGP"
                        )
                        send_telegram_report(report_msg)
                        send_telegram_file(cat_excel_data, "Insurance_Comparison_Categorized.xlsx")

QUOTATION_SHEET_URL = "https://docs.google.com/spreadsheets/d/1-PehpPy5rHuMjO5TVzFWhD8B9z5VqVCou_PMIwnKwH4/edit?pli=1&gid=1268576354#gid=1268576354"
TEMPLATE_DOWNLOAD_URL = "https://docs.google.com/spreadsheets/d/1BBjR_p1ITjf-SNS-gHb7_-WWy22C0ASH/export?format=xlsx"

with tab2:
    st.header("📝 Official Quotation System")

    col_sync, col_temp = st.columns(2)

    with col_sync:
        if st.button("🔄 Sync Rates from Server", use_container_width=True):
            with st.status("📡 Connecting...", expanded=True) as status:
                xl_rates = load_data(QUOTATION_SHEET_URL)
                if xl_rates:
                    st.session_state['xl_rates_data'] = xl_rates
                    status.update(label="✅ Quotation Rates Updated!", state="complete", expanded=False)
                else:
                    status.update(label="❌ Failed to load Quotation Rates", state="error")

    with col_temp:
        st.link_button("📥 Download Employees Template", TEMPLATE_DOWNLOAD_URL,
                       use_container_width=True, type="primary")

    if 'xl_rates_data' in st.session_state:
        xl_rates = st.session_state['xl_rates_data']
        st.markdown("---")

        col_left, col_right = st.columns(2)
        with col_left:
            life_input = st.number_input("Life Value (L.E)", min_value=0.0, value=0.0, key="life_q")
            tax_input  = st.number_input("Fees & Expenses %", min_value=0.0, value=0.0,
                                         step=0.1, key="tax_q") / 100
        with col_right:
            company        = st.selectbox("Select Insurer", xl_rates.sheet_names, key="comp_q")
            employees_file = st.file_uploader("Upload Employees File (Excel)", type=['xlsx'], key="file_q")

        if employees_file:
            if st.button("🚀 Generate Official Quotation", use_container_width=True):
                try:
                    with st.spinner("Processing official quotation and formatting Excel..."):
                        df_people  = pd.read_excel(employees_file)
                        price_list = pd.read_excel(xl_rates, sheet_name=company)
                        price_list.columns = [str(c).strip() for c in price_list.columns]

                        plan_codes_row = price_list.iloc[0]
                        available_cols = [c for c in price_list.columns
                                          if c != 'Plan Name' and 'Unnamed' not in str(c)]
                        code_to_plan_map = {str(plan_codes_row[c]).split('.')[0]: c
                                            for c in available_cols}

                        results     = []
                        plan_counts = {}

                        for _, person in df_people.iterrows():
                            dob          = pd.to_datetime(person['BirthDate'], dayfirst=True, errors='coerce')
                            emp_plan_code = str(person['Plan Code']).split('.')[0].strip()

                            error_found = False
                            if emp_plan_code in code_to_plan_map:
                                p_name   = code_to_plan_map[emp_plan_code]
                                age      = calculate_insurance_age(dob, company_name=company)
                                medical_p = get_price(age, price_list, p_name)
                                plan_counts[p_name] = plan_counts.get(p_name, 0) + 1
                            else:
                                p_name, medical_p, age = "Not Found", 0, 0
                                error_found = True

                            life_used        = life_input
                            current_tax_rate = tax_input
                            if company in ["كاف", "Kaf"]:
                                rel = str(person.get('Releation', '')).strip().upper()
                                if rel == 'D':
                                    life_used        = 0
                                    current_tax_rate = max(tax_input - 0.0104, 0)

                            pre_fees = medical_p + life_used
                            fees_v   = pre_fees * current_tax_rate

                            results.append({
                                'Name': person['Name'], 'Age': age, 'Plan Name': p_name,
                                'Plan Code': emp_plan_code, 'Medical Premium': medical_p,
                                'Life Insurance': life_used, 'Subtotal': pre_fees,
                                'Issuance Fees & Expenses': fees_v,
                                'Grand Total': pre_fees + fees_v,
                                'Is_Error': error_found
                            })

                        output = BytesIO()
                        writer = pd.ExcelWriter(output, engine='xlsxwriter')
                        df_out = pd.DataFrame(results)
                        df_out.drop(columns=['Is_Error']).to_excel(
                            writer, index=False, sheet_name='Quotation', startrow=5)

                        workbook  = writer.book
                        worksheet = writer.sheets['Quotation']

                        header_fmt = workbook.add_format({
                            'bold': True, 'font_size': 16, 'font_color': '#FFFFFF',
                            'bg_color': '#203764', 'align': 'center', 'valign': 'vcenter', 'border': 1
                        })
                        warning_fmt = workbook.add_format({
                            'bold': True, 'font_size': 12, 'font_color': '#FF0000',
                            'align': 'center', 'valign': 'vcenter'
                        })
                        num_fmt = workbook.add_format({
                            'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '#,##0.00'
                        })
                        error_fmt = workbook.add_format({
                            'bg_color': '#FFC7CE', 'border': 1, 'align': 'center', 'num_format': '#,##0.00'
                        })
                        total_fmt = workbook.add_format({
                            'bold': True, 'bg_color': '#FFFF00', 'border': 1, 'align': 'center',
                            'num_format': '#,##0.00'
                        })

                        worksheet.merge_range('A1:I2',
                            'Ahmed Soufy - Corporate Medical Account Manager ©', header_fmt)
                        worksheet.merge_range('A3:I4',
                            'أسعار تجريبية عرضة للزيادة أو النقصان طبقاً لشروط شركة التأمين',
                            warning_fmt)
                        worksheet.protect('123')

                        worksheet.set_column('A:A', 30)
                        worksheet.set_column('B:I', 18)

                        for r_idx, data in enumerate(results):
                            fmt = error_fmt if data['Is_Error'] else num_fmt
                            row_vals = [
                                data['Name'], data['Age'], data['Plan Name'], data['Plan Code'],
                                data['Medical Premium'], data['Life Insurance'], data['Subtotal'],
                                data['Issuance Fees & Expenses'], data['Grand Total']
                            ]
                            for c_idx, val in enumerate(row_vals):
                                worksheet.write(r_idx + 6, c_idx, val, fmt)

                        last_row_data = len(results) + 6
                        for c_idx in [4, 5, 6, 7, 8]:
                            col_letter = chr(65 + c_idx)
                            worksheet.write_formula(
                                last_row_data, c_idx,
                                f"=SUM({col_letter}7:{col_letter}{last_row_data})", total_fmt)

                        stats_fmt      = workbook.add_format({'bold': True, 'italic': True,
                                                              'font_size': 10, 'align': 'left'})
                        footer_copy_fmt = workbook.add_format({'italic': True, 'font_size': 9,
                                                               'align': 'right', 'font_color': '#595959'})

                        stats_row_idx = last_row_data + 2
                        if 0 < len(plan_counts) <= 5:
                            stats_text = "Plans Summary: " + " | ".join(
                                [f"{k} - {v}" for k, v in plan_counts.items()])
                            worksheet.merge_range(
                                f'A{stats_row_idx + 1}:E{stats_row_idx + 1}', stats_text, stats_fmt)

                        footer_text = "Made by: Ahmed Soufy - Corporate Medical Account Manager ©"
                        worksheet.merge_range(
                            f'F{stats_row_idx + 1}:I{stats_row_idx + 1}', footer_text, footer_copy_fmt)

                        for i in range(9):
                            worksheet.set_column(i, i, 20)

                        writer.close()

                        total_premium = sum(item['Grand Total'] for item in results)
                        report_msg = (
                            f"🚀 <b>New Quotation Generated!</b>\n\n"
                            f"🏢 <b>Company:</b> {company}\n"
                            f"👥 <b>Employees:</b> {len(results)}\n"
                            f"💰 <b>Total Premium:</b> {total_premium:,.2f} EGP"
                        )
                        send_telegram_report(report_msg)
                        send_telegram_file(output.getvalue(), f"Quotation_{company}.xlsx")

                        st.success(f"✅ {company} Quotation Successfully Generated ! ✅ ")
                        st.download_button(
                            label="📥 Download Official Excel Quotation",
                            data=output.getvalue(),
                            file_name=f"Quotation_{company}_Official.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                except Exception as e:
                    st.error(f"Error: {e}")
    else:
        st.warning("Please press the 'Sync Rates' button first .. ⚠️ ")


# ════════════════════════════════════════════
# TAB 3 – New Ideas
# ════════════════════════════════════════════
with tab3:
    st.header("💡 New Ideas & Future Updates")
    st.info("This section is currently **Under Development**. "
            "We are working on bringing more automation to your insurance workflow!")

    col_idea1, col_idea2 = st.columns(2)
    with col_idea1:
        st.subheader("🚀 Coming Soon")
        st.write("""
        - Automated policy comparison.
        - Smart recommendations based on client history.
        """)
    with col_idea2:
        st.subheader("📊 Advanced Analytics")
        st.write("""
        - Interactive dashboards for your sales performance.
        - Real-time market share tracking.
        - Client retention alerts.
        """)


st.divider()
st.markdown(
    """
    <div style="text-align: left; font-size: 13px; font-style: italic;">
        Developed with ❤️ by Ahmed Soufy | @ 2026
    </div>
    """,
    unsafe_allow_html=True
)

